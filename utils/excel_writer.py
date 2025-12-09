"""
Excel出力ユーティリティ
DataFrameをExcelに出力し、色分けを行う
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelWriter:
    """Excel出力を行うクラス"""

    # 色定義
    COLOR_GREEN = 'C6EFCE'  # 完全一致（緑）
    COLOR_RED = 'FFC7CE'    # 候補なし・低スコア（赤）
    COLOR_YELLOW = 'FFEB9C'  # 取引先候補1（黄色）
    COLOR_YELLOW_LIGHT = 'FFF9E6'  # 取引先元・候補2-3（薄い黄色）
    COLOR_BLUE = 'DDEBF7'    # 部門候補1（青色）
    COLOR_BLUE_LIGHT = 'F0F6FC'    # 部門元・候補2-3（薄い青色）
    COLOR_WHITE = 'FFFFFF'   # デフォルト（白）

    def __init__(self):
        self.wb = None
        self.ws = None

    def write_to_excel(self, df, output_path, sheet_name='Sheet1'):
        """
        DataFrameをExcelに出力し、色分けを行う

        Args:
            df: 出力するDataFrame
            output_path: 出力先ファイルパス
            sheet_name: シート名

        Returns:
            str: 出力先ファイルパス
        """
        # 新しいワークブックを作成
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name

        # DataFrameをワークシートに書き込み
        columns = list(df.columns)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx, value=value)

                # ヘッダー行のスタイル
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                # データ行で金額列の場合
                elif r_idx > 1 and c_idx <= len(columns):
                    col_name = columns[c_idx - 1]
                    if '金額' in col_name:
                        # 数値の場合、三桁カンマと右揃えを適用
                        if isinstance(value, (int, float)) and pd.notna(value):
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right', vertical='center')

        # 行の高さを1.5倍に設定
        self._set_row_height(df)

        # 列幅を自動調整
        self._auto_fit_columns(df)

        # 色分けを適用
        self._apply_colors(df)

        # 罫線を適用
        self._apply_borders(df)

        # ファイルを保存
        self.wb.save(output_path)

        return output_path

    def _set_row_height(self, df):
        """
        行の高さを1.5倍に設定

        Args:
            df: DataFrame
        """
        # デフォルトの行の高さは15、1.5倍にする
        default_height = 15
        new_height = default_height * 1.5

        # すべての行に適用（ヘッダー含む）
        for row_idx in range(1, len(df) + 2):  # ヘッダー + データ行
            self.ws.row_dimensions[row_idx].height = new_height

    def _auto_fit_columns(self, df):
        """
        列幅を自動調整（日本語対応）

        Args:
            df: DataFrame
        """
        for idx, column in enumerate(df.columns, 1):
            max_width = self._calculate_text_width(str(column))

            # データの最大幅を取得
            for value in df[column]:
                if pd.notna(value):
                    text_width = self._calculate_text_width(str(value))
                    if text_width > max_width:
                        max_width = text_width

            # 最大幅を設定（最大60、最小10）
            adjusted_width = min(max(max_width + 2, 10), 60)
            self.ws.column_dimensions[self._get_column_letter(idx)].width = adjusted_width

    def _calculate_text_width(self, text):
        """
        テキストの表示幅を計算（日本語対応）

        日本語文字は英数字の約2倍の幅を取るため、それを考慮する

        Args:
            text: テキスト

        Returns:
            float: 表示幅
        """
        width = 0
        for char in text:
            # 日本語文字（ひらがな、カタカナ、漢字、全角記号）
            if ord(char) > 127:
                width += 2
            else:
                width += 1
        return width

    def _get_column_letter(self, col_idx):
        """
        列番号をアルファベットに変換

        Args:
            col_idx: 列番号（1始まり）

        Returns:
            str: 列アルファベット（例: 1→A, 27→AA）
        """
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _apply_colors(self, df):
        """
        色分けを適用

        ルール:
        - 完全一致: 緑
        - 不一致: 赤
        - 取引先候補列: 黄色
        - 部門候補列: 青色

        Args:
            df: DataFrame
        """
        columns = list(df.columns)

        # 各行に対して処理
        for row_idx in range(len(df)):
            excel_row = row_idx + 2  # Excelの行番号（ヘッダーが1行目）

            # 取引先のチェック
            if '_取引先完全一致' in columns:
                is_partner_match = df.at[row_idx, '_取引先完全一致']
                has_partner = df.at[row_idx, 'STREAMED元の取引先'] if 'STREAMED元の取引先' in columns else ''

                if pd.notna(has_partner) and has_partner != '':
                    color = self.COLOR_GREEN if is_partner_match else self.COLOR_RED
                    self._color_row(excel_row, columns, color, exclude_patterns=['候補', '_'])

            # 部門のチェック
            if '_部門完全一致' in columns:
                is_dept_match = df.at[row_idx, '_部門完全一致']
                has_dept = df.at[row_idx, 'STREAMED元の部門'] if 'STREAMED元の部門' in columns else ''

                if pd.notna(has_dept) and has_dept != '':
                    # すでに取引先で色が付いていない場合のみ部門の色を適用
                    if '_取引先完全一致' not in columns or pd.isna(df.at[row_idx, 'STREAMED元の取引先']) or df.at[row_idx, 'STREAMED元の取引先'] == '':
                        color = self.COLOR_GREEN if is_dept_match else self.COLOR_RED
                        self._color_row(excel_row, columns, color, exclude_patterns=['候補', '_'])

        # 候補列に色を付ける
        self._color_candidate_columns(df, columns)

        # フラグ列（_で始まる列）を非表示にする
        self._hide_flag_columns(columns)

    def _color_row(self, excel_row, columns, color, exclude_patterns=None):
        """
        行全体に色を付ける

        Args:
            excel_row: Excelの行番号
            columns: 列名リスト
            color: 色コード
            exclude_patterns: 除外するパターンのリスト（列名に含まれる文字列）
        """
        if exclude_patterns is None:
            exclude_patterns = []

        for col_idx, col_name in enumerate(columns, 1):
            # 除外パターンに一致する列はスキップ
            skip = False
            for pattern in exclude_patterns:
                if pattern in col_name:
                    skip = True
                    break

            if not skip:
                cell = self.ws.cell(row=excel_row, column=col_idx)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    def _hide_flag_columns(self, columns):
        """
        フラグ列（_で始まる列）を非表示にする

        Args:
            columns: 列名リスト
        """
        for col_idx, col_name in enumerate(columns, 1):
            if col_name.startswith('_'):
                col_letter = self._get_column_letter(col_idx)
                self.ws.column_dimensions[col_letter].hidden = True

    def _color_candidate_columns(self, df, columns):
        """
        候補列に色を付ける

        Args:
            df: DataFrame
            columns: 列名リスト
        """
        for col_idx, col_name in enumerate(columns, 1):
            # STREAMED元の取引先列: 薄い黄色
            if col_name == 'STREAMED元の取引先':
                self._color_column(col_idx, len(df), self.COLOR_YELLOW_LIGHT)

            # 取引先候補1: 濃い黄色
            elif col_name == '取引先候補1':
                self._color_column(col_idx, len(df), self.COLOR_YELLOW)

            # 取引先候補2-3: 薄い黄色
            elif col_name in ['取引先候補2', '取引先候補3']:
                self._color_column(col_idx, len(df), self.COLOR_YELLOW_LIGHT)

            # STREAMED元の部門列: 薄い青色
            elif col_name == 'STREAMED元の部門':
                self._color_column(col_idx, len(df), self.COLOR_BLUE_LIGHT)

            # 部門候補1: 濃い青色
            elif col_name == '部門候補1':
                self._color_column(col_idx, len(df), self.COLOR_BLUE)

            # 部門候補2-3: 薄い青色
            elif col_name in ['部門候補2', '部門候補3']:
                self._color_column(col_idx, len(df), self.COLOR_BLUE_LIGHT)

    def _color_column(self, col_idx, row_count, color):
        """
        特定の列に色を付ける

        Args:
            col_idx: 列番号（1始まり）
            row_count: 行数
            color: 色コード
        """
        for row_idx in range(2, row_count + 2):  # ヘッダーを除く
            cell = self.ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    def _apply_borders(self, df):
        """
        罫線を適用

        Args:
            df: DataFrame
        """
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        columns = list(df.columns)

        for row_idx, row in enumerate(self.ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border

                # 金額列の場合は右揃えを維持、それ以外は左揃え
                if row_idx > 1 and col_idx <= len(columns):
                    col_name = columns[col_idx - 1]
                    if '金額' in col_name:
                        # 金額列は右揃えを維持
                        if cell.alignment is None or cell.alignment.horizontal != 'right':
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        # それ以外は左揃え
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    # ヘッダー行は左揃え
                    cell.alignment = Alignment(horizontal='left', vertical='center')
